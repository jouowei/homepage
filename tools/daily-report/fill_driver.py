#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""日報表司機欄自動填寫工具。

輸入：一個車輛日報表 .xlsx ＋ 若干個 LINE 出勤群組匯出 .txt
輸出：填好司機欄的完整活頁簿、單獨匯出的「日報表」分頁、對應檢核報告

用法：
    python3 fill_driver.py                 # 讀 inputs/、寫 outputs/
    python3 fill_driver.py 某資料夾
    python3 fill_driver.py --input-dir in --output-dir out --sheet 日報表
"""
import argparse
import json
import os
import re
import sys
from collections import Counter, OrderedDict, defaultdict

try:
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("缺少 openpyxl，請先執行： pip install openpyxl")


# ---------------------------------------------------------------- 文字剖析

HAN = r"一-鿿"
TIMESTAMP = re.compile(r"^\d{1,2}:\d{2}\s+(.*)$")          # LINE 每則訊息的時間前綴
DATE_SEP = re.compile(r"^\d{4}[./-]\d{1,2}[./-]\d{1,2}")   # 「2026.07.03 星期五」分隔列

# 出勤表標題：「7/1出勤」（早班群、夜配群）或整列就是「7/1」（大車群）
HEAD_ATTENDANCE = re.compile(r"(\d{1,2})\s*/\s*(\d{1,2})\s*出勤")
HEAD_BARE_DATE = re.compile(r"(?:^|\s)(\d{1,2})\s*/\s*(\d{1,2})\s*$")

# 名單列：「153吳宗賢」（車號在前）或「葉宏傑9310」（姓名在前）
ENTRY_NUM_FIRST = re.compile(r"^(\d{3,4})\s*([%s]{2,4})$" % HAN)
ENTRY_NUM_FIRST_NOTE = re.compile(r"^(\d{3,4})\s*([%s]{2,3})\S*$" % HAN)   # 後面帶備註
ENTRY_NAME_FIRST = re.compile(r"^([%s]{2,4})\s*(\d{3,4})$" % HAN)

# 名單區塊中要略過的說明列（休假、跟車新人、區域小標……）
SKIP_PREFIXES = ("休", "新人", "台東", "廠休", "更正", "老闆", "支援", "備註")


def read_messages(path):
    """把 LINE 匯出檔切成一則則訊息，並去掉時間與發話者前綴。"""
    messages, current = [], None
    with open(path, encoding="utf-8-sig") as fh:
        for raw in fh:
            line = raw.strip()
            if DATE_SEP.match(line):
                continue
            stamped = TIMESTAMP.match(line)
            if stamped:
                current = [stamped.group(1).strip()]
                messages.append(current)
            else:
                if current is None:
                    current = []
                    messages.append(current)
                current.append(line)
    return messages


def sniff_format(messages):
    """依命中次數判斷這個群組用哪一種名單寫法。"""
    num_first = name_first = 0
    for msg in messages:
        for line in msg:
            if ENTRY_NUM_FIRST.match(line):
                num_first += 1
            elif ENTRY_NAME_FIRST.match(line):
                name_first += 1
    return "num_first" if num_first >= name_first else "name_first"


def match_entry(line, style):
    if style == "name_first":
        m = ENTRY_NAME_FIRST.match(line)
        return (m.group(2), m.group(1)) if m else None
    m = ENTRY_NUM_FIRST.match(line) or ENTRY_NUM_FIRST_NOTE.match(line)
    return (m.group(1), m.group(2)) if m else None


def resolve_year(month, base_year, base_month):
    """名單只寫月/日；跨年時往前後各推一年。"""
    if month - base_month > 6:
        return base_year - 1
    if base_month - month > 6:
        return base_year + 1
    return base_year


def apply_plate_fixes(date, plate, name, fixes, report):
    """套用「某人的車號被打錯」的修正規則，回傳更正後的車號代碼。"""
    for fix in fixes:
        if fix["司機"] != name or str(fix["誤植"]) != plate:
            continue
        if str(fix.get("起", "")) and date < str(fix["起"]):
            continue
        if str(fix.get("迄", "")) and date > str(fix["迄"]):
            continue
        report.append("  車號修正 %s %s：%s → %s" % (date, name, plate, fix["正確"]))
        return str(fix["正確"])
    return plate


def parse_roster(path, group, base_year, base_month, aliases, plate_fixes, report):
    """回傳 {yyyymmdd: [(車號代碼, 司機), ...]}；同日重複公布時只留最後一次。"""
    messages = read_messages(path)
    style = sniff_format(messages)
    head = HEAD_BARE_DATE if style == "name_first" else HEAD_ATTENDANCE
    report.append("讀取 %s（%s，格式：%s）" % (os.path.basename(path), group,
                                          "姓名+車號" if style == "name_first" else "車號+姓名"))

    rosters = OrderedDict()
    for msg in messages:
        head_at = None
        for i, line in enumerate(msg):
            m = head.search(line)
            if m:
                head_at, month, day = i, int(m.group(1)), int(m.group(2))
                break
        if head_at is None:
            continue
        date = "%04d%02d%02d" % (resolve_year(month, base_year, base_month), month, day)

        rows = []
        for line in msg[head_at + 1:]:
            if not line or line.startswith(SKIP_PREFIXES):
                continue
            hit = match_entry(line, style)
            if hit:
                plate, name = hit
                name = aliases.get(name, name)
                rows.append((apply_plate_fixes(date, plate, name, plate_fixes, report), name))
            else:
                report.append("  未解析 %d/%d：%r" % (month, day, line))
        if not rows:
            continue

        if date in rosters and dict(rosters[date]) != dict(rows):
            report.append("  %s 出勤表有更正，採用最後一次公布" % date)
        rosters[date] = rows
    return rosters


# ---------------------------------------------------------------- 車號對應

def plate_code(plate):
    """取車號中的數字段：063-Q8 -> 063、KPB-9310 -> 9310、128-9B -> 128。"""
    for token in str(plate).split("-"):
        if token.isdigit():
            return token
    return None


def build_code_map(plates, report):
    code_to_plate, collisions = {}, defaultdict(list)
    for plate in plates:
        code = plate_code(plate)
        if not code:
            report.append("警告：車號 %r 取不出數字段，無法對應" % plate)
            continue
        collisions[code].append(plate)
        code_to_plate[code] = plate
    for code, group in collisions.items():
        if len(group) > 1:
            report.append("警告：代碼 %s 對應到多台車 %s，已採用 %s" % (code, group, code_to_plate[code]))
    return code_to_plate


# ---------------------------------------------------------------- 主流程

def load_settings(input_dir, report):
    """選用的 設定.json，四個區塊都可以不寫。"""
    path = os.path.join(input_dir, "設定.json")
    if not os.path.exists(path):
        return {}, [], [], []
    with open(path, encoding="utf-8-sig") as fh:
        cfg = json.load(fh)
    aliases = cfg.get("aliases", {})
    fixes = cfg.get("車號修正", [])
    fixed = cfg.get("固定司機", [])
    overrides = cfg.get("overrides", [])
    report.append("套用 設定.json：別名 %d、車號修正 %d、固定司機 %d、人工修正 %d"
                  % (len(aliases), len(fixes), len(fixed), len(overrides)))
    return aliases, fixes, fixed, overrides


def main():
    here = os.path.dirname(os.path.abspath(__file__))
    ap = argparse.ArgumentParser(description="依 LINE 出勤名單填寫日報表司機欄")
    ap.add_argument("input_dir", nargs="?", default=os.path.join(here, "inputs"),
                    help="放 1 個 .xlsx 與若干 .txt 的資料夾（預設 inputs/）")
    ap.add_argument("--output-dir", default=os.path.join(here, "outputs"))
    ap.add_argument("--sheet", default="日報表")
    ap.add_argument("--driver-col", type=int, default=1, help="司機欄（預設 A=1）")
    ap.add_argument("--plate-col", type=int, default=2, help="車號欄（預設 B=2）")
    ap.add_argument("--date-col", type=int, default=3, help="日期欄（預設 C=3）")
    ap.add_argument("--work-col", type=int, default=4, help="上班時間欄，用來檢核（預設 D=4）")
    args = ap.parse_args()

    input_dir, out_dir = args.input_dir, args.output_dir
    if not os.path.isdir(input_dir):
        sys.exit("找不到輸入資料夾：%s" % input_dir)
    os.makedirs(out_dir, exist_ok=True)

    books = sorted(f for f in os.listdir(input_dir)
                   if f.lower().endswith(".xlsx") and not f.startswith("~$"))
    texts = sorted(f for f in os.listdir(input_dir) if f.lower().endswith(".txt"))
    if len(books) != 1:
        sys.exit("inputs 資料夾需要剛好 1 個 .xlsx，目前有 %d 個" % len(books))
    if not texts:
        sys.exit("inputs 資料夾裡沒有任何 .txt 出勤名單")

    report = ["=== 日報表司機對應檢核報告 ===", ""]
    aliases, plate_fixes, fixed_drivers, overrides = load_settings(input_dir, report)

    book_path = os.path.join(input_dir, books[0])
    report.append("日報表：%s" % books[0])
    wb = openpyxl.load_workbook(book_path)
    if args.sheet not in wb.sheetnames:
        sys.exit("活頁簿裡沒有「%s」分頁，現有分頁：%s" % (args.sheet, wb.sheetnames))
    ws = wb[args.sheet]

    # 以日報表本身的日期決定年月
    dates = [str(ws.cell(r, args.date_col).value or "")
             for r in range(2, ws.max_row + 1)]
    dates = [d for d in dates if len(d) == 8 and d.isdigit()]
    if not dates:
        sys.exit("「%s」的日期欄讀不到 yyyymmdd 格式的資料" % args.sheet)
    base = Counter(d[:6] for d in dates).most_common(1)[0][0]
    base_year, base_month = int(base[:4]), int(base[4:6])
    report.append("報表月份：%s 年 %s 月" % (base_year, base_month))
    report.append("")

    # 逐群組讀出勤名單；後讀到的群組不會覆蓋先前群組已寫入的車號
    roster = OrderedDict()
    for name in texts:
        group = os.path.splitext(name)[0]
        for date, rows in parse_roster(os.path.join(input_dir, name), group, base_year,
                                       base_month, aliases, plate_fixes, report).items():
            for code, driver in rows:
                key = (date, code)
                if key in roster and roster[key][0] != driver:
                    report.append("  衝突 %s 車號%s：%s / %s（保留 %s）"
                                  % (date, code, roster[key][0], driver, roster[key][0]))
                    continue
                roster[key] = (driver, group)

    for item in overrides:
        key = (str(item["日期"]), str(item["車號"]))
        if item.get("動作") == "移除":
            roster.pop(key, None)
        else:
            roster[key] = (aliases.get(item["司機"], item["司機"]), "人工修正")
        report.append("人工修正：%s %s %s %s"
                      % (item.get("動作", "新增"), item["日期"], item["車號"],
                         item.get("司機", ""), ))

    report.append("")
    report.append("出勤名單共 %d 筆（日期 × 車號）" % len(roster))

    # 寫回司機欄
    plates = [ws.cell(r, args.plate_col).value for r in range(2, ws.max_row + 1)]
    plate_of_code = build_code_map({p for p in plates if p}, report)
    code_of_plate = {plate: code for code, plate in plate_of_code.items()}

    # 同一人同一天出現在兩台車上，通常代表某群組漏了更正，列出來讓人工確認
    same_day = defaultdict(list)
    for (date, code), (driver, _) in roster.items():
        same_day[(date, driver)].append(code)
    doubled = {k: v for k, v in same_day.items() if len(v) > 1}
    report += ["", "--- 同一天掛到兩台以上車輛的司機（請人工確認） ---"]
    for (date, driver), codes in sorted(doubled.items()):
        report.append("  %s %s：%s" % (date, driver, "、".join(sorted(codes))))
    if not doubled:
        report.append("  （無）")

    # 固定司機：整台車包給一個人、名單上從來不會出現的，用車號代碼指定
    fixed_by_code = {}
    for item in fixed_drivers:
        code = plate_code(item["車號"]) or str(item["車號"])
        fixed_by_code[code] = aliases.get(item["司機"], item["司機"])

    filled = 0
    fixed_filled = Counter()
    used = set()
    driver_days = Counter()
    unmatched_activity = Counter()
    for r in range(2, ws.max_row + 1):
        plate = ws.cell(r, args.plate_col).value
        date = ws.cell(r, args.date_col).value
        if not plate or not date:
            continue
        code = code_of_plate.get(plate, "")
        hit = roster.get((str(date), code))
        if hit:
            ws.cell(r, args.driver_col).value = hit[0]
            driver_days[hit[0]] += 1
            used.add((str(date), code))
            filled += 1
        elif code in fixed_by_code and ws.cell(r, args.work_col).value:
            # 固定司機沒有出勤名單可查，所以只填有行駛紀錄的日子
            driver = fixed_by_code[code]
            ws.cell(r, args.driver_col).value = driver
            driver_days[driver] += 1
            fixed_filled[plate] += 1
            filled += 1
        elif ws.cell(r, args.work_col).value:
            unmatched_activity[plate] += 1

    total = ws.max_row - 1
    report += [
        "",
        "資料列 %d，已填司機 %d（%.1f%%），空白 %d" % (total, filled, filled * 100.0 / total, total - filled),
        "司機共 %d 位" % len(driver_days),
    ]
    if fixed_by_code:
        report += ["", "--- 固定司機（依設定填入，只填有行駛紀錄的日子） ---"]
        for plate, days in sorted(fixed_filled.items()):
            report.append("  %-10s %s %d 天" % (plate, fixed_by_code[code_of_plate[plate]], days))
        for code, driver in fixed_by_code.items():
            if code not in plate_of_code:
                report.append("  設定的車號 %s（%s）不在日報表中" % (code, driver))
    report += ["", "--- 有行駛紀錄但出勤名單沒列到的車輛 ---"]
    for plate, days in sorted(unmatched_activity.items(), key=lambda x: (-x[1], x[0])):
        report.append("  %-10s %d 天" % (plate, days))

    leftovers = [k for k in roster if k not in used]
    report += ["", "--- 出勤名單有、但日報表查無該車該日的紀錄 ---"]
    for date, code in sorted(leftovers):
        report.append("  %s 車號%s（%s）%s" % (date, code, roster[(date, code)][0],
                                            "" if code in plate_of_code else " ← 此車號不在日報表中"))
    if not leftovers:
        report.append("  （無）")

    report += ["", "--- 各司機出勤天數 ---"]
    for driver, days in driver_days.most_common():
        report.append("  %-6s %d 天" % (driver, days))

    # 三個輸出
    stem = os.path.splitext(books[0])[0]
    full_path = os.path.join(out_dir, "%s_已填司機.xlsx" % stem)
    wb.save(full_path)

    sheet_path = os.path.join(out_dir, "%s_%s.xlsx" % (args.sheet, base))
    out = openpyxl.Workbook()
    o = out.active
    o.title = args.sheet
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
        o.append(list(row))
    for cell in o[1]:
        cell.font = Font(name="Arial", bold=True)
    for col in range(1, ws.max_column + 1):
        width = max(len(str(o.cell(r, col).value or "")) for r in range(1, min(o.max_row, 200) + 1))
        o.column_dimensions[get_column_letter(col)].width = min(max(width + 2, 10), 24)
    o.freeze_panes = "A2"
    out.save(sheet_path)

    report_path = os.path.join(out_dir, "對應檢核報告.txt")
    with open(report_path, "w", encoding="utf-8") as fh:
        fh.write("\n".join(report) + "\n")

    print("\n".join(report))
    print()
    for path in (full_path, sheet_path, report_path):
        print("已輸出：%s" % path)


if __name__ == "__main__":
    main()
